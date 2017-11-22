#-------------------------------------------------------------------------------
# Name:        Line_K
# Purpose:     Calcualte line K values for selected Areas
#
# Author:      kristjan.vilgo
#
# Created:     04.02.2017
# Copyright:   (c) kristjan.vilgo 2017
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import os, sys
from openpyxl import Workbook
from openpyxl import load_workbook
from operator import sub
#from openpyxl.utils import (_get_column_letter)
from datetime import datetime   #Debug
startTime = datetime.now()      #Debug

#LineRatings vajab:
import pandas as pd
import numpy as np
from LineRatings import get_ratings_and_list_of_buses

print "Start"                   #Debug

#PSSE model

PSSE_model_name="BaasMudel"

#PSSE_model_name="base_case_winter"


PSSE_model_sav=r"""{}.sav""".format(PSSE_model_name)

#Areas, Observable Areas and Outages configuration file

outages_configuration_sheet=            "Outages"
areas_configuration_sheet=              "Areas"
observable_areas_configuration_sheet=   "Observable Areas"

configuration_excel="{}_conf_1.xlsx".format(PSSE_model_name)



#Powerflow settings

powerflow_normal=       2
powerflow_N1=           3
powerflow_area_scale=   3


powerflow_settings=[[0,0,0,1,0,0,99,0], #0 (Full Newton-Raphson/Lock taps/Lock shunts)
                    [0,0,0,1,0,0,-1,0], #1 (Full Newton-Raphson, Ignore reactive power, lock taps and steps.)
                    [1,0,1,1,2,0,99,0], #2
                    [0,0,0,0,0,0,99,0]] #3

##OPTIONS(1) tap adjustment flag (use tap adjustment option setting by default).
##    OPTIONS(1) = 0 disable.
##    OPTIONS(1) = 1 enable stepping adjustment.
##    OPTIONS(1) = 2 enable direct adjustment.
##OPTIONS(2) area interchange adjustment flag (use area interchange adjustment option setting by default).
##    OPTIONS(2) = 0 disable.
##    OPTIONS(2) = 1 enable using tie line flows only in calculating
##    area interchange.
##    OPTIONS(2) = 2 enable using tie line flows and loads in calculating area interchange.
##OPTIONS(3) phase shift adjustment flag (use phase shift adjustment option setting by default).
##    OPTIONS(3) = 0 disable.
##    OPTIONS(3) = 1 enable.
##OPTIONS(4) dc tap adjustment flag (use dc tap adjustment option setting by default).
##    OPTIONS(4) = 0 disable.
##    OPTIONS(4) = 1 enable.
##OPTIONS(5) switched shunt adjustment flag (use switched shunt adjustment option setting by default).
##    OPTIONS(5) = 0 disable.
##    OPTIONS(5) = 1 enable.
##    OPTIONS(5) = 2 enable continuous mode, disable discrete mode.
##OPTIONS(6) flat start flag (0 by default).
##    OPTIONS(6) = 0 do not flat start.
##    OPTIONS(6) = 1 flat start.
##OPTIONS(7) var limit flag (99 by default).
##    OPTIONS(7) = 0 apply var limits immediately.
##    OPTIONS(7) = >0 apply var limits on iteration n (or sooner if mismatch gets small).
##    OPTIONS(7) = -1 ignore var limits.
##OPTIONS(8) non-divergent solution flag (use non-divergent solution option setting by default).
##    OPTIONS(8) = 0 disable.
##    OPTIONS(8) = 1 enable.

solution_states = ["Met convergence tolerance",
                    "Iteration limit exceeded",
                    "Blown up (only when non-divergent option disabled)",
                    "Terminated by non-divergent option",
                    "Terminated by console interrupt",
                    "Singular Jacobian matrix or voltage of 0.0 detected",
                    "Inertial power flow dispatch error (INLF)",
                    "OPF solution met convergence tolerance (NOPF)",
                    "Solution not attempted."]


#Reporting citeria

min_kV=200
max_kV=750



#Functions

def excel_sheet_to_list(excel_file,excel_sheet_name):

    excel_workbook = load_workbook(excel_file)

    excel_sheet=excel_workbook.get_sheet_by_name(excel_sheet_name)

    data = []

    for row in excel_sheet.iter_rows():
        data_row = []
        for cell in row:
            data_row.extend([str(cell.value)]) #Retunrs all values as string, reduces type errors
        data.append(data_row)

    data.pop(0)

    return data


def production(area): #Production of an area


    ierr, rarray = psspy.aareareal(-1, 2, "O_PGEN")
    ierr, iarray = psspy.aareaint(-1, 2, "NUMBER")
    area_pos=iarray[0].index(area)
    print rarray[0][area_pos]
    return rarray[0][area_pos]

def swich_branch(branch,status):

    if branch[0]!="-":
        psspy.branch_chng(int(branch[1]),int(branch[2]),branch[3],[status,_i,_i,_i,_i,_i],[_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f,_f])

def gen_scale(gen,area):

    SID=7
    psspy.bsys(SID,0,[ 0.4, 750.],1,[int(area)],0,[],0,[],0,[])
    psspy.scal_2(SID,0,1,[0,0,0,0,0],[0.0,0.0,0.0,0.0,0.0,0.0,0.0])
    psspy.scal_2(SID,1,2,[_i,3,0,1,0],[ 0.0,gen,0.0,0.0,0.0,0.0,0.0]) #incremental change
    reset_subsys(SID) #Area selection back to default

def reset_subsys(SubSystem_ID):

    psspy.bsys(SubSystem_ID,0,[ 0.4, 750.],0,[],0,[],0,[],0,[]) #Area selection back to default

def get_lines_data(observable_areas,min_kV,max_kV,text_data_string,meta_data_string,data_string,comment_for_header):

    #----Gunnar:
    #SID = 6 #Hardcoded unused SID of system, if in use for other process, this must be changed


    #observable_areas_IDs=map(int,observable_areas) #Converts from string to int

    #psspy.bsys(SID,1,[min_kV,max_kV],len(observable_areas),observable_areas_IDs,0,[],0,[],0,[]) #Sub System definiton

    SID=8 #Use subsystem of unique buses

    #----/Gunnar

    ##SID = a negative value, to instruct the API to assume a subsystem containing all buses in the working case.
    ##SID = a valid bus subsystem identifier.

    OWNER = 1 #to use bus ownership.
    ##OWNER = 2 #to use branch ownership.

    TIES = 1 #for each end of interior subsystem branches only.
    ##TIES = 2 #for the subsystem bus end of tie branches only.
    ##TIES = 3 #for the non-subsystem bus end of tie branches only.
    ##TIES = 4 #for each end of tie branches only.
    ##TIES = 5 #for each end of interior subsystem branches and the subsystem bus end of tie branches.
    ##TIES = 6 #for each end of interior subsystem branches and tie branches.

    ##FLAG = 1 for only in-service branches.
    FLAG = 2 #for all branches.

    ENTRY = 1 #for single entry (each branch once).
    ##ENTRY = 2 #for double entry (each branch in both directions).

    ierr, line_meta_data = psspy.abrnint(SID,OWNER,TIES,FLAG,ENTRY,meta_data_string) #Get metadata of lines ["FROMNUMBER","TONUMBER","STATUS"]

    ierr, line_text_data = psspy.abrnchar(SID,OWNER,TIES,FLAG,ENTRY,text_data_string) #Get text data of lines ["ID", "FROMNAME", "TONAME"]

    ierr, line_data = psspy.abrnreal(SID,OWNER,TIES,FLAG,ENTRY,data_string) #Get data of lines ["AMPS"]

    #Gunnar: reset_subsys(SID) #Area selection back to default

    #List of column headers
    headers=[]
    headers.extend(text_data_string)
    headers.extend(meta_data_string)
    headers.extend(data_string)

    for header_position,header in enumerate(headers):
        headers[header_position]=header+comment_for_header


    #List of data columns
    data=[]
    data.extend(line_text_data)
    data.extend(line_meta_data)
    data.extend(line_data)

    for column_number,column in enumerate(data):
        column.insert(0,headers[column_number]) #Inserts headers to first row of a column


    return data

def get_data_column(column_name_in_header,data_header,data_columns):

    column_ID=data_header.index(column_name_in_header)

    column_data=data_columns[column_ID]

    return column_data

def create_excel_report(list_of_sheets_data,report_name): #List of sheets must be postition [[Sheet_name,List_of_sheet_data_rows],]

    wb = Workbook()

    for n, sheet in enumerate(list_of_sheets_data):
        excel_sheet=wb.create_sheet(sheet[0],n)

        for row in sheet[1]:
            excel_sheet.append(row)
    wb.save(report_name)

def create_excel_report_pd(list_of_sheets_data,report_name):
    writer = pd.ExcelWriter(report_name, engine='xlsxwriter')
    for sheet in list_of_sheets_data:
        sheet[1].to_excel(writer,sheet_name=sheet[0])


def return_data_with_rates(index,col):
    try:
        #print(line_rates.loc[index])
        #print(type(line_rates.loc[index]))
        return line_rates.loc[index].to_frame()
    except:
        return pd.Series(index=col,name=index).to_frame()

def tuple_int(tup):
    lst=map(int,list(tup))
    return tuple(lst)



#        excel_sheet.auto_filter.ref= "A:{}".format(_get_column_letter(len(sheet[1][0])))
    #Definition of outages
outages=excel_sheet_to_list(configuration_excel,outages_configuration_sheet)

    #Definition of Areas
areas=excel_sheet_to_list(configuration_excel,areas_configuration_sheet)

    #Definition of observable Areas - Lines within these areas + tielines will be monitored and reported
observable_areas=excel_sheet_to_list(configuration_excel,observable_areas_configuration_sheet)


# Activation of PSSE
PSSE_LOCATION = r"C:\Program Files (x86)\PTI\PSSE33\PSSBIN"
sys.path.append(PSSE_LOCATION)
os.environ['PATH'] = os.environ['PATH'] + ';' +  PSSE_LOCATION
import psspy

_f = psspy.getdefaultreal()
_i = psspy.getdefaultint()
_s= psspy.getdefaultchar()

psspy.psseinit(0)

psspy.case(PSSE_model_sav)

#Start of process

psspy.fnsl(powerflow_settings[powerflow_normal]) #lahendus (Full Newton-Raphson/Lock taps/Lock shunts)

#Get inital data for areas

area_parameters=["LOAD","GEN","LOSS","INT"]

list_of_area_data=[]

for area in areas:
    area_row=[]
    area_row.append(area[0])
    for parameter in area_parameters:
        error,result=psspy.ardat(int(area[1]),parameter)
        area_row.extend([int(result.real)])
    list_of_area_data.append(area_row)

area_parameters.insert(0,"AREA")

list_of_area_data.insert(0,area_parameters)


#-------Gunnar:
#Get Line ratings and list of buses:
buses, line_rates=get_ratings_and_list_of_buses()

line_rates=line_rates.drop_duplicates()
#Define our buses as subsystem:
num_buses=len(buses)
SID=8 #Do not use elsewhere (unique)

ierr=psspy.bsys(SID, 0, [ 0.4, 750.], 0, [], num_buses, buses, 0, [], 0, [])
'''
print(psspy.abuscount(SID,2))
print(psspy.abrncount(SID, 1, 1, 2, 1))
print(psspy.abrncount(SID, 1, 2, 2, 1))
print(psspy.abrncount(SID, 1, 3, 2, 1))
'''

#------//Gunnar


AREAS=[] #List of area numbers
for area in observable_areas:
    AREAS.append(area[1])

#Get all line flows

    #Iterate trough all outages, N-1 and areas

data=[]

for outage in outages:

    psspy.case(PSSE_model_sav)

    swich_branch(outage,0)

    psspy.fnsl(powerflow_settings[powerflow_normal])

    contingencies=[]

    contingencies.extend(outages) #Presumption is that outage list contains also contingency list

    if outage[0] != "-":

        contingencies.remove(outage) #Remove current outage from contingency list, unless it is the base case

    for contingency in contingencies:

        #print outage[0], contingency[0] #Debug

        columns=[]

        swich_branch(contingency,0)

        psspy.fnsl(powerflow_settings[powerflow_N1]) #lahendus (Full Newton-Raphson/Lock taps/Lock shunts)


            #Line flows (outage+N-1)

        columns.extend(get_lines_data(AREAS,min_kV,max_kV,["ID", "FROMNAME", "TONAME"],["FROMNUMBER","TONUMBER"],["AMPS"],""))#State with N-1

            #Outage column to report

        outage_column=["Outage"]+[outage[0]]*(len(columns[0])-1)

        columns.extend([outage_column])

            #Contingency column to report

        contingency_column=["N-1"]+[contingency[0]]*(len(columns[0])-1)

        columns.extend([contingency_column])

            #Debug, get solution state

        solution_state = psspy.solved()

        #print solution_state, contingency #Debug, prints solution state

        solution_state_column= ["Solved"]+[solution_states[solution_state]]*(len(columns[0])-1)

        columns.extend([solution_state_column])


            #Line flows when areas are scaled

        for area in areas:

            gen_scale(1,area[1]) #Change generation in Area by +1MW

            psspy.fnsl(powerflow_settings[powerflow_area_scale]) #lahendus (Full Newton-Raphson/Lock taps/Lock shunts)

            columns.extend(get_lines_data(AREAS,min_kV,max_kV,[],[],["AMPS"],"_"+area[0])) #Change in Area

            gen_scale(-1,area[1]) #Change generation in Area by -1MW to initial state

        swich_branch(contingency,1)

        rows=(zip(*columns))

        if contingency[0] != "-" or outage[0] != "-" : #Remove header row when it is not the base case

            rows.pop(0) #Will be used later for K calcualtion

        data.extend(rows)

#Calcualte K values



K_data_rows=[] #seperate list for K calculation

K_data_rows.extend(data) #Makes a copy of Line data list

data_header=K_data_rows.pop(0) #Gives us header row for indexses and removes header from data

K_data_columns=zip(*K_data_rows) #Convert rows to columns


base_state_column_name="AMPS"

base_state_column=get_data_column(base_state_column_name,data_header,K_data_columns)


K_value_columns=[]
for area in areas:

    area_column_name=base_state_column_name+"_"+area[0]

    area_column=get_data_column(area_column_name,data_header,K_data_columns)


    K_value_column = map(sub, area_column, base_state_column)

    K_value_column.insert(0,area[0])

    K_value_columns.append(K_value_column)

line_data_columns=zip(*data)

list_of_columns_to_be_added_to_K_data_report=["FROMNAME", "TONAME","Outage","N-1","AMPS"]#"FROMNUMBER","TONUMBER"

for item in list_of_columns_to_be_added_to_K_data_report:

    data_column=get_data_column(item,data_header,line_data_columns)
    K_value_columns.append(data_column)

K_value_rows=zip(*K_value_columns)

#Gunnar:
columns_data=data[0]
data_pd=pd.DataFrame(data=data,columns=columns_data)
data_pd=data_pd[data_pd.ID!="ID"]
ibus=data_pd.apply(lambda row: min([row['FROMNUMBER'], row['TONUMBER']]), axis=1)
jbus=data_pd.apply(lambda row: max([row['FROMNUMBER'], row['TONUMBER']]), axis=1)
data_pd=pd.concat([data_pd,ibus,jbus],axis=1)
data_pd = data_pd.rename(columns={0: 'ibus', 1: 'jbus'})
del data_pd["FROMNUMBER"]
del data_pd["TONUMBER"]
data_pd[["ibus","jbus","ID"]]=data_pd[["ibus","jbus","ID"]].astype(int)
data_pd=data_pd.set_index(['ibus',"jbus","ID"])

line_rates[["ibus","jbus","ickt"]]=line_rates[["ibus","jbus","ickt"]].astype(int)
ibus1=line_rates.apply(lambda row: min([row['ibus'], row['jbus']]), axis=1)
jbus1=line_rates.apply(lambda row: max([row['ibus'], row['jbus']]), axis=1)
line_rates=pd.concat([line_rates,ibus1,jbus1],axis=1)
del line_rates["ibus"]
del line_rates["jbus"]
line_rates = line_rates.rename(columns={0: 'ibus', 1: 'jbus',"ickt":"ID"})
line_rates[["ibus","jbus","ID"]]=line_rates[["ibus","jbus","ID"]].astype(int)
line_rates=line_rates.set_index(["ibus","jbus","ID"])

#print data_pd.head()
#print line_rates.head()
cols=line_rates.columns
#result = pd.concat([data_pd, line_rates], axis=1, join='inner')
#count=0

for index, row in data_pd.iterrows():
    index=tuple_int(index)
    r=row.to_frame()
    #r=r.transpose()
    #data_w_rates=return_data_with_rates(index,cols).append(r)
    #data_w_rates=pd.concat([r,return_data_with_rates(index,cols)],axis=0)
    data_w_rates=r.append(return_data_with_rates(index,cols))
    data_w_rates=data_w_rates.transpose()

    if 'dat' in globals():
        dat=pd.concat([dat,data_w_rates],axis=0)
    else:
        dat=data_w_rates
    """
    count=count+1
    if count==10:
        import sys
        sys.exit("Error message")
    """

#df_rates=pd.DataFrame(list_of_series,list_of_indexes)

#print dat.head()

columns_data=data[0]
data_pd=pd.DataFrame(data=data,columns=columns_data)
data_pd=data_pd[data_pd.ID!="ID"]
K_headers=K_value_rows.pop(0)
#K_indexes=map(list, zip(*K_value_rows)).pop(0)
area_headers=list_of_area_data.pop(0)
#area_indexes=map(list, zip(*list_of_area_data)).pop(0)
K_value_rows_pd = pd.DataFrame(K_value_rows, columns=K_headers)
list_of_area_data_pd= pd.DataFrame(list_of_area_data, columns=area_headers)
#print result.head()

#Write results to excel
"""
sheets_data=[   ["Line_Data",data],
                ["K_Data",K_value_rows],
                ["Area_Data",list_of_area_data]   ]
"""
sheets_data_pd=[["Line_Data",dat],["K_Data",K_value_rows_pd],["Area_Data",list_of_area_data_pd]]

report_name="{}_lines_K_report.xlsx".format(PSSE_model_name)

#create_excel_report(sheets_data,report_name)

create_excel_report_pd(sheets_data_pd,report_name)

#Print execution time

run_duration=datetime.now() - startTime

print "Done - [{}s]".format(run_duration.seconds)


###Debug - prints report header row
##header=[]
##header.extend(zip(*columns)[0])
##print header










